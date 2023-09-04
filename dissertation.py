# -*- coding: utf-8 -*-
"""
Created on Thu Aug 17 15:21:22 2023

@author: JUNGHYUN HONG
"""

"""
Regarding github data export
https://computer-science-student.tistory.com/297 

When you want to import a file from github , import it in the format below.
https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path} 
"""

"""
This code sample shows Prebuilt Layout operations with the Azure Form Recognizer client library. 
The async versions of the samples require Python 3.6 or later.

To learn more, please visit the documentation - Quickstart: Form Recognizer Python client library SDKs
https://docs.microsoft.com/en-us/azure/applied-ai-services/form-recognizer/quickstarts/try-v3-python-sdk
"""

"""
Remember to remove the key from your code when you're done, and never post it publicly. For production, use
secure methods to store and access your credentials. For more information, see 
https://docs.microsoft.com/en-us/azure/cognitive-services/cognitive-services-security?tabs=command-line%2Ccsharp#environment-variables-and-application-configuration
"""

import json
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import FormRecognizerClient , FormTrainingClient
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import sys
from openpyxl import Workbook


"""
Don't erase it.
"""
# document_analysis_client = DocumentAnalysisClient(
#     endpoint=endpoint, credential=AzureKeyCredential(key)
# )
# document_analysis_client = FormRecognizerClient(endpoint,AzureKeyCredential(key))





endpoint = "https://eastasia.api.cognitive.microsoft.com/"
key = "41c6a81c79ab43369eb5901219ffaed3"

# sample document
formUrl = r"C:\Users\Rainbow Brain\Desktop\논문\An overview of biodegradable packaging in food industry.pdf"


"""
Create form_recognizer Client by FormRecognizerClient  
"""
form_recognizer_client = FormRecognizerClient(endpoint,AzureKeyCredential(key))


""" 
Pull in a pdf file from github and create poller  
When you want to import a file from github , import it in the format below.
                                     ↓↓↓↓↓
https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path}  
"""
poller = form_recognizer_client.begin_recognize_content_from_url("https://raw.githubusercontent.com/hjh5333/honghong/main/An overview of biodegradable packaging in food industry.pdf")
# poller = document_analysis_client.begin_analyze_document_from_url("prebuilt-read", "https://raw.githubusercontent.com/hjh5333/honghong/main/An overview of biodegradable packaging in food industry.pdf")
result = poller.result()



print('EXCEL FILE OPEN')
excel_file = openpyxl.load_workbook(r"C:\Users\Rainbow Brain\Desktop\python_prc\theisTable\tableSave.xlsx")
excel_sheet = excel_file['Sheet1']
print('EXCEL FILE OPEN COMPLETE')

""" An index equal to the number of pages is created in the 'result'(list)  """
for idx , page in enumerate(result): 
    if idx == 0:
        excelRowCount = 1
    else:
        excelRowCount += 5
    """ The number of indexes in page.tables varies depending on how many tables there are in the page.  """
    for table in page.tables: 
        print("ColumnCount : {0}".format(table.column_count))
        print('row Count {0}'.format(table.row_count))
        columnCount = table.column_count + 1
        rowCount = table.row_count
        excelColumnCount = 1
        for cell  in table.cells: 
            excel_sheet.cell( row = cell.row_index + 1 , column = cell.column_index + 1 ).value = cell.text
            if excelColumnCount == columnCount :
                excelRowCount += 1
                excelColumnCount = 1
            else :
                excelColumnCount += 1
            print('Cell value : {0}'.format(cell.text))
            # print("row: {0} : column : {1}".format(cell.row_index,cell.column_index))
            # print('Confidence Score : {0}'.format(cell.confidence))

excel_file.save(r"C:\Users\Rainbow Brain\Desktop\python_prc\theisTable\tableSave.xlsx")
print('excelfile 저장완료')
exit()



""" 나중에 사용해 볼 부분 """
# for idx, style in enumerate(result.styles):
#     print(
#         "Document contains {} content".format(
#          "handwritten" if style.is_handwritten else "no handwritten"
#         )
#     )

# for page in result.pages:
#     for line_idx, line in enumerate(page.lines):
#         print(
#          "...Line # {} has text content '{}'".format(
#         line_idx,
#         line.content.encode("utf-8")
#         )
#     )

#     for selection_mark in page.selection_marks:
#         print(
#          "...Selection mark is '{}' and has a confidence of {}".format(
#          selection_mark.state,
#          selection_mark.confidence
#          )
#     )

# print("==================================================================================")
# print("==================================================================================")

# for table_idx, table in enumerate(result.tables):
#     print(
#         "Table # {} has {} rows and {} columns".format(
#         table_idx, table.row_count, table.column_count
#         )
#     )
        
#     for cell in table.cells:
#         print(
#             "...Cell[{}][{}] has content '{}'".format(
#             cell.row_index,
#             cell.column_index,
#             cell.content.encode("utf-8"),
#             )
#         )

print("----------------------------------------")

